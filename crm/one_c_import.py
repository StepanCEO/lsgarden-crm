from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import BinaryIO

from django.db import transaction
from openpyxl import load_workbook

from .demo_data import DEMO_PRODUCT_NAMES
from .models import Product


NOMENCLATURE_SHEET = 'номенклатура Т'


@dataclass
class ImportStats:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    deleted_demo: int = 0

    @property
    def total(self) -> int:
        return self.created + self.updated


def _clean_text(value) -> str:
    return str(value or '').strip()


def _as_decimal(value) -> Decimal:
    raw = _clean_text(value).replace(' ', '').replace(',', '.')
    if not raw:
        return Decimal('0')
    try:
        return Decimal(raw)
    except InvalidOperation:
        return Decimal('0')


def _as_int(value) -> int:
    text = _clean_text(value)
    if not text:
        return 0
    try:
        return int(float(text))
    except ValueError:
        return 0


def _product_kind(category: str, name: str) -> str:
    haystack = f'{category} {name}'.lower()
    if any(token in haystack for token in ('горш', 'кашпо', 'букет', 'удобр', 'аксессуар', 'грунт')):
        return Product.ProductKind.OTHER
    return Product.ProductKind.PLANT


def _product_status(stock: int) -> str:
    if stock <= 0:
        return Product.StockStatus.CRITICAL
    if stock <= 5:
        return Product.StockStatus.LOW
    return Product.StockStatus.OK


def _open_workbook(source: str | Path | BinaryIO):
    if hasattr(source, 'seek'):
        source.seek(0)
    return load_workbook(source, data_only=True)


def _iter_nomenclature_rows(source: str | Path | BinaryIO):
    workbook = _open_workbook(source)
    if NOMENCLATURE_SHEET not in workbook.sheetnames:
        raise ValueError(f'В файле не найден лист "{NOMENCLATURE_SHEET}"')

    sheet = workbook[NOMENCLATURE_SHEET]
    for row in sheet.iter_rows(values_only=True):
        product_id = row[0] if len(row) > 0 else None
        name = _clean_text(row[1] if len(row) > 1 else '')
        if product_id is None or not name:
            continue
        try:
            numeric_id = int(product_id)
        except (TypeError, ValueError):
            continue

        category = _clean_text(row[2] if len(row) > 2 else '') or 'Комнатные растения'
        price = _as_decimal(row[4] if len(row) > 4 else '')
        stock = _as_int(row[5] if len(row) > 5 else 0)
        yield {
            'sku': str(numeric_id),
            'name': name,
            'parent': category,
            'kind': _product_kind(category, name),
            'stock': stock,
            'reserve': 0,
            'price': price,
            'in_production': 0,
            'status': _product_status(stock),
        }


@transaction.atomic
def import_nomenclature(source: str | Path | BinaryIO) -> ImportStats:
    stats = ImportStats()
    stats.deleted_demo = Product.objects.filter(name__in=DEMO_PRODUCT_NAMES).count()
    if stats.deleted_demo:
        Product.objects.filter(name__in=DEMO_PRODUCT_NAMES).delete()

    existing_by_sku = {product.sku: product for product in Product.objects.all()}
    to_create: list[Product] = []
    to_update: list[Product] = []

    for payload in _iter_nomenclature_rows(source):
        product = existing_by_sku.get(payload['sku'])
        if product is None:
            to_create.append(Product(**payload))
            stats.created += 1
            continue

        changed = False
        for field, value in payload.items():
            if getattr(product, field) != value:
                setattr(product, field, value)
                changed = True
        if changed:
            to_update.append(product)
            stats.updated += 1
        else:
            stats.skipped += 1

    if to_create:
        Product.objects.bulk_create(to_create, batch_size=500)
    if to_update:
        Product.objects.bulk_update(
            to_update,
            ['name', 'parent', 'kind', 'stock', 'reserve', 'price', 'in_production', 'status', 'updated_at'],
            batch_size=500,
        )

    return stats
